# -*- coding: utf-8 -*-
# Copyright 2020-present Lenovo
# Copyright 2015-2023 Lenovo
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import logging
import os
import string

from django.conf import settings
from django.utils.translation import ugettext as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..utils import get_local_timezone
from .etc import (
    CLUSTER_DAILY_DATA_STYLE, CLUSTER_DAILY_END_LINE_DATA,
    CLUSTER_DAILY_END_LINE_STYLE, CLUSTER_DAILY_HEAD,
    CLUSTER_MONTHLY_DATA_STYLE, CLUSTER_MONTHLY_END_LINE_DATA,
    CLUSTER_MONTHLY_HEAD, RESOURCE_UNIT, USER_DAILY_END_LINE_DATA,
    USER_DAILY_END_LINE_STYLE, USER_DAILY_JOB_DATA_STYLE,
    USER_DAILY_STORAGE_DATA_STYLE, USER_MONTHLY_DATA_STYLE,
    USER_MONTHLY_END_LINE_DATA, USER_MONTHLY_END_LINE_STYLE, USER_MONTHLY_HEAD,
    get_create_info, user_daily_report_head,
)

logger = logging.getLogger(__name__)


class BillingReportExport(object):
    export_path = settings.ACCOUNTING.BILLING_DIR
    encoding = 'utf-8'

    def __init__(self, data, doctype, start_time, end_time, username,
                 create_time, billgroup, actual_total_cost, storage_data,
                 job_data, bill_date, gres, bill_name):
        self.data = data
        self.bill_name = bill_name
        self.username = username
        self.create_time = create_time.astimezone(
            tz=get_local_timezone()
        ).strftime('%F %T')
        self.billgroup = billgroup
        self.storage_data = storage_data
        self.start_time = start_time.astimezone(
            tz=get_local_timezone()
        ).strftime('%F %T')
        self.end_time = end_time.astimezone(
            tz=get_local_timezone()
        ).strftime('%F %T')
        self.job_data = job_data
        self.report_export = getattr(self, 'export_' + doctype)
        self.bill_date = bill_date.astimezone(
            tz=get_local_timezone()
        ).strftime('%Y-%m-%d')
        self.actual_total_cost = actual_total_cost
        self.gres = gres
        self.bill_month = bill_date.astimezone(
            tz=get_local_timezone()
        ).strftime('%Y-%m')
        self.title_cell = 'A1'
        if self.job_data:
            self.job_data.sort(key=lambda x: x['record_id'])
        self.cost_unit_format = self.convert_unit_format()

    def convert_unit_format(self):
        cost_unit_format = '#,##0.00'
        unit = settings.ACCOUNTING.BILLING.UNIT.split(';', 1)
        if len(unit) <= 1:
            return unit[0].strip() + cost_unit_format
        else:
            if len(unit[1]) == 0:
                return unit[0].strip() + cost_unit_format
            else:
                return unit[0].strip() + cost_unit_format + unit[1].strip()

    def export_xls(self):
        export_billing = {
            'User_Daily_Bills': self.daily_user_report,
            'Daily_Summary_Bills': self.daily_cluster_report,
            'User_Monthly_Bills': self.monthly_user_report,
            'Monthly_Summary_Bills': self.monthly_cluster_report
        }
        username = '_{0}'.format(self.username) if self.username else ''
        if 'Daily' in self.bill_name:
            prefix = self.bill_name+username+'_{0}_{1}'.format(
                self.bill_date, settings.ACCOUNTING.BILLING.LANGUAGE)
        else:
            prefix = self.bill_name + username + '_{0}_{1}'.format(
                self.bill_month, settings.ACCOUNTING.BILLING.LANGUAGE)
        suffix = '.xlsx'
        filename = prefix + suffix
        export_path = os.path.join(settings.ACCOUNTING.BILLING_DIR, filename)

        export_billing[self.bill_name](Workbook(), export_path)
        return filename

    def monthly_user_report(self, wb, export_path):
        ws = wb.active
        if self.data:
            self.data.sort(key=lambda x: x['billing_date'])
        data_len = len(self.data)
        # set create info
        info_dict = get_create_info(
            self.bill_name, user=self.username,
            start_time=self.start_time, end_time=self.end_time,
            create_time=self.create_time, total=self.actual_total_cost
        )
        self.set_create_info(ws, info_dict)

        # set month bill head
        self.set_bill_head(ws, USER_MONTHLY_HEAD['user_monthly_head'])

        # set month bill data
        total_runtime = 0
        data_begin_line = ws.max_row+1
        for row in range(data_begin_line, data_len+data_begin_line):
            data_list = list(self.data[row-data_begin_line].values())
            runtime = self.set_data_style(
                ws, data_list, USER_MONTHLY_DATA_STYLE, row)
            total_runtime += runtime

        # set month bill end line
        self.set_end_line(
            ws, USER_MONTHLY_END_LINE_DATA, USER_MONTHLY_END_LINE_STYLE,
            ws.max_row+1, data_len, total_runtime=total_runtime)

        # flush monthly bill style
        self.flush_title_style(ws, USER_MONTHLY_HEAD.get('title', ''))

        wb.save(export_path)

    def monthly_cluster_report(self, wb, export_path):
        ws = wb.active
        if self.data:
            self.data.sort(key=lambda x: x['username'])
        data_len = len(self.data)

        # set create info
        info_dict = get_create_info(
            self.bill_name, bill_month=self.bill_month,
            start_time=self.start_time, end_time=self.end_time,
            create_time=self.create_time, total=self.actual_total_cost
        )
        self.set_create_info(ws, info_dict)

        # set cluster month bill head
        self.set_bill_head(ws, CLUSTER_MONTHLY_HEAD['cluster_monthly_head'])

        # set cluster month bill data
        data_begin_line = ws.max_row + 1
        total_runtime = 0
        for row in range(data_begin_line, data_len + data_begin_line):
            data_list = list(self.data[row-data_begin_line].values())
            runtime = self.set_data_style(
                ws, data_list, CLUSTER_MONTHLY_DATA_STYLE, row)
            total_runtime += runtime

        # set cluster month bill end line
        self.set_end_line(
            ws, CLUSTER_MONTHLY_END_LINE_DATA, CLUSTER_MONTHLY_DATA_STYLE,
            ws.max_row + 1, data_len, total_runtime=total_runtime)

        # flush cluster monthly bill style
        self.flush_title_style(ws, CLUSTER_MONTHLY_HEAD.get('title', ''))

        wb.save(export_path)

    def daily_cluster_report(self, wb, export_path):
        ws = wb.active
        if self.data:
            self.data.sort(key=lambda x: x['username'])
        data_len = len(self.data)
        # set create info
        info_dict = get_create_info(
            self.bill_name, bill_date=self.bill_date,
            create_time=self.create_time, total=self.actual_total_cost
        )
        self.set_create_info(ws, info_dict)

        # set daily bill head
        self.set_bill_head(ws, CLUSTER_DAILY_HEAD['cluster_daily_head'])

        # set daily bill data
        total_runtime = 0
        data_begin_line = ws.max_row + 1
        for row in range(data_begin_line, data_len+data_begin_line):
            data_list = list(self.data[row - data_begin_line].values())
            runtime = self.set_data_style(
                ws, data_list, CLUSTER_DAILY_DATA_STYLE, row)
            total_runtime += runtime

        # set daily bill end line
        self.set_end_line(
            ws, CLUSTER_DAILY_END_LINE_DATA, CLUSTER_DAILY_END_LINE_STYLE,
            ws.max_row + 1, data_len, total_runtime=total_runtime)

        # flush daily bill style
        self.flush_title_style(ws, CLUSTER_DAILY_HEAD.get('title', ''))

        wb.save(export_path)

    def daily_user_report(self, wb, export_path):
        ws = wb.active
        s_len = len(self.storage_data)
        j_len = len(self.job_data)

        # set create info
        info_dict = get_create_info(
            self.bill_name, user=self.username,
            bill_group=self.billgroup, bill_date=self.bill_date,
            create_time=self.create_time, total=self.actual_total_cost
        )
        self.set_create_info(ws, info_dict)

        # write storage head
        head_dict = user_daily_report_head(s_len if s_len == 0 else s_len - 1)
        self.set_user_daily_head(ws, head_dict['storage_head'])

        # write storage bill data
        data_begin_line = ws.max_row + 1
        for row in range(data_begin_line, s_len+data_begin_line):
            data_list = list(self.storage_data[row - data_begin_line].values())
            self.set_data_style(
                ws, data_list, USER_DAILY_STORAGE_DATA_STYLE, row)
            # storage bill merge cell from A column to C column
            ws.merge_cells('A{0}:C{0}'.format(row))

        # write storage end line
        end_row = s_len + data_begin_line
        self.set_end_line(
            ws, USER_DAILY_END_LINE_DATA['storage_end'],
            USER_DAILY_STORAGE_DATA_STYLE, end_row, s_len)

        # write job head
        self.set_user_daily_head(ws, head_dict['job_head'])

        # write job bill data
        total_runtime = 0
        data_begin_line = ws.max_row + 1
        for row in range(data_begin_line, data_begin_line+j_len):
            data_list = list(self.job_data[row-data_begin_line].values())
            runtime = self.set_data_style(
                ws, data_list, USER_DAILY_JOB_DATA_STYLE, row)
            total_runtime += runtime

        # write job end line
        self.set_end_line(
            ws, USER_DAILY_END_LINE_DATA['job_end'], USER_DAILY_END_LINE_STYLE,
            data_begin_line + j_len, j_len,
            total_runtime=total_runtime)

        # flush bill-head style
        self.flush_title_style(ws, head_dict.get('title', ''))

        wb.save(export_path)

    def flush_title_style(self, ws, title_value):
        self.set_col_width(ws, string.ascii_uppercase[0:ws.max_column], 15)
        ws[self.title_cell] = \
            settings.ACCOUNTING.BILLING.TITLE + _(title_value)
        self.add_style(ws[self.title_cell], font_size=25, has_border=False)
        title_cell = '{0}:{1}1'.format(
            self.title_cell,
            string.ascii_uppercase[ws.max_column - 1]
        )
        ws.merge_cells(title_cell)

    def set_end_line(
            self, ws, end_data, end_style, end_row, data_len, total_runtime=0):
        for index, cell_dict in enumerate(end_data):
            ws[cell_dict['start_cell'].format(row=end_row)] = \
                _(cell_dict['value'].format(
                    start_row=end_row-1 if data_len == 0 else end_row-data_len,
                    end_row=end_row-1, total_runtime=total_runtime))
            self.add_style(
                ws[cell_dict['start_cell'].format(row=end_row)],
                **list(end_style[index].values())[0](
                    'FFFFFF00')
            )
            if cell_dict['start_cell'] != cell_dict['end_cell']:
                ws.merge_cells(
                    cell_dict['start_cell'].format(row=end_row) + ':' +
                    cell_dict['end_cell'].format(row=end_row))

    def set_data_style(self, ws, data_list, data_style, row):
        data_list = self._data_sort(data_list)
        runtime = 0
        for index, style_dict in enumerate(data_style):
            for cell, style in style_dict.items():
                write_style = style('FFFFFF')
                write_style['bold'] = False
                if 'data_format' in write_style:
                    runtime = data_list[index]
                ws[cell.format(row=row)] = data_list[index]
                self.add_style(ws[cell.format(row=row)], **write_style)
        return runtime

    def user_daily_bill_head_update(self, ws, cell_dict, sub_cell):
        unit = settings.ACCOUNTING.BILLING.UNIT + ' / '
        if cell_dict['title'] in RESOURCE_UNIT:
            ws[cell_dict['sub_cell'][0]['start_cell']] = unit + \
                _(cell_dict['sub_cell'][0]['title'])
            self.add_style(ws[sub_cell['start_cell']])
            ws[cell_dict['sub_cell'][1]['start_cell']] = \
                _(cell_dict['sub_cell'][1]['title']).format(
                    unit=_(RESOURCE_UNIT[cell_dict['title']]))
            self.add_style(ws[sub_cell['start_cell']])
        else:
            for g_dict in self.gres:
                if g_dict['display_name'] == cell_dict['title']:
                    ws[cell_dict['sub_cell'][0]['start_cell']] = \
                        unit + g_dict['unit'] + \
                        _(cell_dict['sub_cell'][0]['title'])
                    self.add_style(ws[cell_dict['sub_cell'][0]['start_cell']])
                    ws[cell_dict['sub_cell'][1]['start_cell']] = \
                        _(cell_dict['sub_cell'][1]['title']).format(
                            unit=g_dict['unit'])
                    self.add_style(ws[cell_dict['sub_cell'][1]['start_cell']])

    def set_user_daily_head(self, ws, head_list):
        for cell_dict in head_list:
            ws[cell_dict['start_cell']] = _(cell_dict['title'])
            self.add_style(ws[cell_dict['start_cell']])
            if cell_dict['start_cell'] != cell_dict['end_cell']:
                ws.merge_cells(
                    cell_dict['start_cell'] + ':' + cell_dict['end_cell']
                )
            if 'sub_cell' in cell_dict:
                for sub_cell in cell_dict['sub_cell']:
                    ws[sub_cell['start_cell']] = _(sub_cell['title'])
                    self.add_style(ws[sub_cell['start_cell']])
                    self.user_daily_bill_head_update(ws, cell_dict, sub_cell)

    def bill_head_unit_update(self, ws, cell_dict):
        """
        self.gres  Format:
        [
            {'display_name': 'GPU', 'code': 'gpu', 'unit': 'Card'},
            {'display_name': 'FPGA', 'code': 'fpga', 'unit': 'Card'},
            ......
        ]
        Params: cell_dict  Format:
        {
            "title": "GPU",
            "start_cell": "E6",
            "end_cell": "F6",
            "sub_cell": [
                {
                    "title": "Django.accounting.bill.Hour",
                    "start_cell": "E7",
                    "end_cell": "E7"
                },
                {
                    "title": "Django.accounting.bill.Cost",
                    "start_cell": "F7",
                    "end_cell": "F7"
                },
                ..
            ]
        }
        """
        for g_dict in self.gres:
            if g_dict['display_name'] == cell_dict['title']:
                for index, cell in enumerate(cell_dict['sub_cell']):
                    if cell['title'] == 'Django.accounting.bill.Hour':
                        unit_convert_cell = cell_dict['sub_cell'].pop(index)
                        ws[unit_convert_cell['start_cell']] = \
                            g_dict['unit'] + _(unit_convert_cell['title'])
                        self.add_style(ws[unit_convert_cell['start_cell']])

    def set_bill_head(self, ws, head_list):
        import copy
        for cell_dict in copy.deepcopy(head_list):
            ws[cell_dict['start_cell']] = _(cell_dict['title'])
            if 'sub_cell' in cell_dict:
                self.bill_head_unit_update(ws, cell_dict)
                for sub_cell in cell_dict['sub_cell']:
                    ws[sub_cell['start_cell']] = _(sub_cell['title'])
                    self.add_style(ws[sub_cell['start_cell']])
            self.add_style(ws[cell_dict['start_cell']])
            if cell_dict['start_cell'] != cell_dict['end_cell']:
                ws.merge_cells(
                    cell_dict['start_cell'] + ':' + cell_dict['end_cell']
                )

    def set_create_info(self, ws, info_dict):
        for cell, value in info_dict['title'].items():
            ws[cell] = _(value)
            self.add_style(ws[cell], hori='left', has_border=False)

        for cell, value in info_dict['value'].items():
            ws[cell] = value
            self.add_style(ws[cell], hori='right', has_border=False)
        ws[info_dict['currency_cell']].number_format = self.cost_unit_format

    def add_style(self, cell, **kwargs):
        border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000')
        )
        font = Font(
            name='Arial',
            size=kwargs.get('font_size', 10),
            bold=kwargs.get('bold', True)
        )
        alig = Alignment(
            horizontal=kwargs.get('hori', 'center'),
            vertical='center'
        )
        fill = PatternFill(
            fill_type='solid',
            start_color=kwargs.get('fill_color', 'FFF0F0F0'),
            end_color=kwargs.get('fill_color', 'FFF0F0F0')
        )
        cell.font = font
        cell.alignment = alig
        if kwargs.get('curr_style', False):
            cell.number_format = self.cost_unit_format
        if kwargs.get('keep_decimal', False):
            cell.number_format = '0.00'
        if kwargs.get('data_format', False):
            m, s = divmod(int(cell.value), 60)
            h, m = divmod(m, 60)
            cell.value = "%d:%02d:%02d" % (h, m, s)
        if kwargs.get('has_border', True):
            cell.border = border
            cell.fill = fill

    @staticmethod
    def _data_sort(data_list):
        order_key = ['cost', 'count', 'charge_rate']
        for index, value in enumerate(data_list):
            if isinstance(value, dict):
                gres_dict = data_list.pop(index)
                for key in order_key:
                    if key in gres_dict:
                        data_list.insert(index, gres_dict[key])
        return data_list

    @staticmethod
    def set_col_width(ws, col_list, width):
        for col in col_list:
            ws.column_dimensions[col].width = width
